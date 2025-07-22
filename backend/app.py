import re
from flask import Flask, request, send_file, jsonify
import openpyxl
import tempfile
import os
import csv
from io import BytesIO, StringIO

from flask_cors import CORS

app = Flask(__name__)

# CORS : autorise localhost et ton IP serveur
CORS(app, origins=[
    "http://localhost:5173",
    "http://127.0.0.1:5173",
    "http://91.98.29.232"
])


# Utilitaires pour injection (à adapter selon ta logique métier)
def inject_data(wb, mois, annee, csv_files):
    # print(f"[DEBUG] Début de l'injection pour {mois} {annee}")

    def clean_float(val):
        if not val:
            return 0
        val = re.sub(r"[^0-9,.-]", "", val)  # garde chiffres, virgule, point, moins
        val = val.replace(",", ".")
        try:
            return float(val)
        except Exception:
            return 0

    # Recherche de la feuille année
    ws = None
    for sheet in wb.worksheets:
        if sheet.title == str(annee):
            ws = sheet
            break
    if ws is None:
        raise Exception(f"Feuille {annee} non trouvée")

    # print(f"[DEBUG] Feuille trouvée : {ws.title}")

    # Recherche de la ligne de début du mois
    start_row = None
    for row in ws.iter_rows(min_col=1, max_col=1):
        if row[0].value == mois:
            start_row = row[0].row + 3
            break
    if start_row is None:
        raise Exception(f"Mois {mois} non trouvé")

    # print(f"[DEBUG] Ligne de début trouvée : {start_row}")

    # Détermination du nombre de jours
    mois31 = ["JANVIER", "MARS", "MAI", "JUILLET", "AOUT", "OCTOBRE", "DECEMBRE"]
    if mois in mois31:
        nb_jours = 31
    elif mois == "FEVRIER":
        nb_jours = 29 if int(annee) % 4 == 0 else 28
    else:
        nb_jours = 30

    # print(f"[DEBUG] Nombre de jours pour {mois} {annee} : {nb_jours}")

    jours_a_sauter = []
    if mois in ["JANVIER", "MAI"]:
        jours_a_sauter = [0]
    elif mois == "DECEMBRE":
        jours_a_sauter = [24]

    # print(f"[DEBUG] Jours à sauter : {jours_a_sauter}")

    # Jours fériés non travaillés (1er janvier, 1er mai, 25 décembre)
    jours_feries = []
    if mois == "JANVIER":
        jours_feries.append(0)  # 1er janvier (index 0)
    if mois == "MAI":
        jours_feries.append(0)  # 1er mai (index 0)
    if mois == "DECEMBRE":
        jours_feries.append(24)  # 25 décembre (index 24)

    csv_idx = 0
    for jour in range(nb_jours):
        # print(f"[DEBUG] Traitement du jour {jour + 1}")
        if jour in jours_feries:
            # print(
            #     f"[DEBUG] Jour férié non travaillé (jour {jour + 1}), aucune feuille de caisse consommée."
            # )
            continue
        if csv_idx >= len(csv_files):
            # print(
            #     f"[DEBUG] Plus de fichiers CSV à traiter (csv_idx={csv_idx}), on arrête."
            # )
            break
        csv_file = csv_files[csv_idx]
        csv_file.stream.seek(0)
        # Lecture du CSV avec le séparateur ;
        reader = csv.reader(
            StringIO(csv_file.stream.read().decode("utf-8")), delimiter=";"
        )
        # print(f"[DEBUG] Lecture de {csv_file.filename}")
        # print(f"[DEBUG] Lecture de {reader}")
        rows = list(reader)
        # print(f"[DEBUG] rows = {rows}")
        # print(f"[DEBUG] Nombre de lignes lues : {len(rows)}")

        valRJ = 0
        valRL = 0
        for i in range(0, min(31, len(rows))):
            # print(f"[DEBUG] Traitement de la ligne {i} pour le jour {jour + 1}")
            row = rows[i]
            type_val = row[1] if len(row) > 1 else None
            # print(f"[DEBUG] Type de valeur : {type_val} (à la ligne {i})")
            d_val = clean_float(row[3]) if len(row) > 3 else 0
            c_val = clean_float(row[2]) if len(row) > 2 else 0

            excel_row = start_row + jour
            if type_val == "Espèces":
                # print(f"[DEBUG] Especes à la ligne {excel_row} : {d_val}€")
                ws[f"B{excel_row}"] = d_val
            if type_val == "Chèque":
                # print(f"[DEBUG] Chèque à la ligne {excel_row} : {d_val}€")
                ws[f"C{excel_row}"] = d_val
            if type_val == "CB":
                # print(f"[DEBUG] CB à la ligne {excel_row} : {d_val}€")
                ws[f"D{excel_row}"] = d_val
            if type_val == "Gain tirage et sport":
                # print(
                #     f"[DEBUG] Gain tirage et sport à la ligne {excel_row} : {d_val}€ (ajouté à valRL)"
                # )
                valRL += d_val
            if type_val == "Gain grattage":
                # print(
                #     f"[DEBUG] Gain grattage à la ligne {excel_row} : {d_val}€ (ajouté à valRJ)"
                # )
                valRJ += d_val
            if type_val == "Especes POINT VERT":
                # print(
                #     f"[DEBUG] Especes POINT VERT à la ligne {excel_row} : D={d_val}€, C={c_val}€"
                # )
                ws[f"G{excel_row}"] = d_val
                ws[f"H{excel_row}"] = c_val
            if type_val == "Avoir":
                # print(f"[DEBUG] Avoir à la ligne {excel_row} : {d_val}€")
                ws[f"K{excel_row}"] = d_val
            if type_val == "REMB. JEU":
                # print(
                #     f"[DEBUG] REMB. JEU à la ligne {excel_row} : {d_val}€ (ajouté à valRJ)"
                # )
                valRJ += d_val
            if type_val == "REMB. LOTO":
                # print(
                #     f"[DEBUG] REMB. LOTO à la ligne {excel_row} : {d_val}€ (ajouté à valRL)"
                # )
                valRL += d_val
            if type_val == "Mise en compte":
                # print(f"[DEBUG] Mise en compte à la ligne {excel_row} : {d_val}€")
                ws[f"J{excel_row}"] = d_val
            if type_val == "Paiement facture":
                # print(f"[DEBUG] Paiement facture à la ligne {excel_row} : C={c_val}€")
                ws[f"K{excel_row}"] = c_val
        ws[f"F{start_row + jour}"] = valRL
        ws[f"E{start_row + jour}"] = valRJ
        csv_idx += 1
    return wb


@app.route("/traiter-compta", methods=["POST"])
def traiter_compta():
    excel_file = request.files["excel"]
    csv_files = request.files.getlist("csvs")
    mois = request.form["mois"]
    annee = request.form["annee"]
    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        excel_file.save(tmp.name)
        wb = openpyxl.load_workbook(tmp.name)
        wb = inject_data(wb, mois, annee, csv_files)
        wb.save(tmp.name)
        tmp.seek(0)
        return send_file(
            tmp.name, as_attachment=True, download_name=f"Compta-{mois}-{annee}.xlsx"
        )


@app.route("/status", methods=["GET"])
def status():
    return "ok", 200
    

if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    debug = os.environ.get("FLASK_DEBUG", "1") == "1"
    # Host 0.0.0.0 pour serveur, localhost pour dev
    host = "0.0.0.0" if os.environ.get("RUN_ON_SERVER", "0") == "1" else "127.0.0.1"
    app.run(host=host, port=port, debug=debug)
