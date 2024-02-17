import openpyxl as pyxl
import glob
import tkinter as tk
import tkinter.font
from tkinter import filedialog

class ComptaAuto:
    tablfile = ""         ## valeur de stockage du lien du paquet de copie
    paquetfile= ""
    mois = ""
    annee = 0
    lignedebut = 0
    version= "Version 1.2"
    nbJours = 0

    def AffilierTableur(self,file ):       ## deff d'affiliation de correctionfile Appellée apres pression sur le bouton
        self.tablfile = file

    def fillJour(self, lignecourrante ,worksheet, file):
        wbjour = pyxl.load_workbook(file)  ## wb du jour
        wsjour = wbjour.active  ## sheet de la fiche de caisse du jour
        valRJ = 0  ## valeur remboursement Jeu
        valRL = 0  ## valeur remboursement Loto
        for ligne in range(
                30):  ## parcours des lignes dans la feuille de caisse puis check des valeurs des cases dans la colone B pour aller chercher les valeurs correspondantes et les mettre dans le tbleur final
            case = 'B' + str(ligne + 1)  ## memoire pour savoir quelle ligne on se situ

            if wsjour[case].value == "Espèces":  ##test si correspond à espece
                casevaleur = 'D' + str(
                    ligne + 1)  ## case dans laquellle on recupere le nombre de la fiche de caisse
                worksheet['B' + lignecourrante].value = wsjour[
                    casevaleur].value  ## mise dans la valeur du tableau de compta

            if wsjour[case].value == "Chèque":
                casevaleur = 'D' + str(ligne + 1)
                worksheet['C' + lignecourrante].value = wsjour[casevaleur].value

            if  wsjour[case].value == "CB":
                casevaleur = 'D' + str(ligne + 1)
                worksheet['D' + lignecourrante].value = wsjour[casevaleur].value


            if wsjour[case].value == "Gain tirage et sport":

                casevaleur = 'D' + str(ligne + 1)
                valRL = wsjour[casevaleur].value + valRL
                print(" test gaintirage loto DANS TIRAGE :")
                print(wsjour[casevaleur].value)

            if wsjour[case].value == "Gain grattage":
                casevaleur = 'D' + str(ligne + 1)
                valRJ = wsjour[casevaleur].value + valRJ

            if wsjour[case].value == "Especes POINT VERT":
                casevaleur = 'D' + str(ligne + 1)
                casenb = 'C' + str(ligne + 1)
                worksheet['I' + lignecourrante].value = wsjour[casevaleur].value
                worksheet['J' + lignecourrante].value = wsjour[casenb].value

            if wsjour[case].value == "Avoir":
                casevaleur = 'D' + str(ligne + 1)
                worksheet['M' + lignecourrante].value = wsjour[casevaleur].value

            if wsjour[case].value == "REMB. JEU":
                casevaleur = 'D' + str(ligne + 1)
                valRJ = valRJ + wsjour[casevaleur].value

            if wsjour[case].value == "REMB. LOTO":
                casevaleur = 'D' + str(ligne + 1)
                print(" TEST RB LOTO AVANT ADD:")
                print(wsjour[casevaleur].value)
                valRL = valRL + wsjour[casevaleur].value
                print(" DANS TEST REMB LOTO " + str(valRL))


            if wsjour[case].value == "Mise en compte":
                casevaleur = 'D' + str(ligne + 1)
                worksheet['O' + lignecourrante].value = wsjour[casevaleur].value

            if wsjour[case].value == "Paiement facture":
                casevaleur = 'C' + str(ligne + 1)
                worksheet['P' + lignecourrante].value = wsjour[casevaleur].value

        print("FIN ADD:" + str(valRL))
        worksheet['H' + lignecourrante].value = valRL  ## remboursemenent loto car 2 cas et impossible de faire .value + .value
        worksheet['G' + lignecourrante].value = valRJ


    def AffilierDoss(self, file):       ## def d'affiliation de paquetfile Apellée apres pression dur le bouton
        self.paquetfile = file

    def AffilierMA(self, mois, annee):  ## same pour nbquestion
        self.mois = mois
        self.annee = int(annee)

    def Execution(self,):
        TablAnnee = self.tablfile           ## variable

        file_list = glob.glob(self.paquetfile + '/*.*') ## pour récuprer le ficher de toutes les feuilles de caisses
        print(file_list)
        wb = pyxl.load_workbook(TablAnnee)          ## recupération de la feuille de compta
        file_list.sort()

        for sheet in wb:            ## recherche de la bonne feuille de la bonne année
            if sheet.title == str(self.annee):           ##si le nom de la feuille corrsepond à l'année
                ws = sheet                          ##stockage de la sheet
        nombre = 1      ## pour stocker et creer la case de départ
        colone = 'A'        ## pour faire les cases

        while (nombre < 467) :                      ## recherche du mois dans le tbleur de compta
            casedate = colone + str(nombre)
            if ws[casedate].value == self.mois:
                self.lignedebut = nombre + 3            ## + 3 car la case de juillet est 3 case au dessus de la premiere case à remplir
            nombre = nombre + 1

        if self.mois == "JANVIER" or self.mois == "MARS" or self.mois == "MAI" or self.mois == "JUILLET" or self.mois == "AOUT" or self.mois == "OCTOBRE" or self.mois == "DECEMBRE":
            self.nbjours = 31
        elif self.mois == "FEVRIER":
            if self.annee%4 == 0:
                self.nbjours = 29
            else:
                self.nbjours = 28
        else : self.nbjours =30

        if self.mois == "JANVIER":
            for jour in range(self.nbjours):  ## parcours des feuilles de caisses par jour
                if jour == 1:           ## si on est le premier janvier pas de caisse a faire
                    print()
                else:
                    self.fillJour( str(self.lignedebut + jour), ws, file_list[jour-1])

        elif self.mois == "MAI":
            for jour in range(self.nbjours):  ## parcours des feuilles de caisses par jour
                if jour == 0:
                    print()
                else:
                    self.fillJour( str(self.lignedebut + jour), ws, file_list[jour-1])

        elif self.mois == "DECEMBRE":
            for jour in range(self.nbjours):  ## parcours des feuilles de caisses par jour
                if jour == 24:
                    print()
                elif jour <24:
                    self.fillJour( str(self.lignedebut + jour), ws, file_list[jour])
                elif jour >24:
                    self.fillJour( str(self.lignedebut + jour), ws, file_list[jour])
        else:
            for jour in range(self.nbjours):         ## parcours des feuilles de caisses par jour
                self.fillJour( str(self.lignedebut + jour), ws, file_list[jour])
        print("chaussure :) ")    ## j'ai retrouvé mes chaussures!!
        wb.save(TablAnnee)          ##save de la feuille de compta




class Application(tk.Tk):
    def __init__(self):
        tk.Tk.__init__(self)
        self.geometry('400x400')
        self.resizable(width=0, height=0)
        self.iconbitmap("Logo valid.ico")
        self.title("Compta Automatique")
        self.Compta = ComptaAuto()
        self.creer_widgets(self.Compta)
        self.police = tkinter.font.Font(size=5)

    def creer_widgets(self, Compta : ComptaAuto):
        self.canva1 = tk.Canvas(self, width=400, height=400)  ## canva principal qui fait toute la fenetre
        self.canva1.create_rectangle(20, 20, 380, 50, outline="black")  ## rectangle du haut, pour mettre titre
        self.canva1.create_text(200, 35, text="Compta Automatique Maman") ## Titre dans le rectangle
        self.canva1.place(x=0, y=0) ## placer le canva
        self.canva1.create_oval(360, 60, 390, 90, fill='red')   ## voyant check tableur
        self.canva1.create_oval(360,105,390,135, fill = 'red')  ## voyant check Paquet de copie

        self.canva2 = tk.Canvas(self, width=200,height = 75)
        self.canva2.create_text(100,25, text="MOIS ANNEE",font= tkinter.font.Font(size=12)) ## Titre dans le rectangle
        self.canva2.place(x=200,y=225)

        OptionList = ["JANVIER","FEVRIER","MARS","AVRIL","MAI","JUIN","JUILLET","AOUT","SEPTEMBRE","OCTOBRE","NOVEMBRE", "DECEMBRE"]
        self.variable = tk.StringVar(self)
        self.variable.set(OptionList[0])
        self.opt = tk.OptionMenu(self, self.variable, *OptionList)
        self.opt.config(width=10, font=('Helvetica', 12))
        self.opt.place(x=20,y=180)

        OptionList2 = [2018, 2019, 2020, 2021, 2022, 2023, 2024, 2025, 2026]
        self.variable2 = tk.StringVar(self)
        self.variable2.set(OptionList2[4])
        self.opt2 = tk.OptionMenu(self, self.variable2, *OptionList2)
        self.opt2.config(width=10, font=('Helvetica', 12))
        self.opt2.place(x=200, y=180)

        self.bouton_askCompta = tk.Button( text="Donner le tableur", command=lambda: Compta.AffilierTableur(self.AskComptaFile())).place(x=20, y=70)
        self.bouton_askDoss = tk.Button( text="Donner le paquet de feuille de caisse", command=lambda: Compta.AffilierDoss(self.AskPackFile())).place(x=20, y=110)
        self.canva1.create_text(125, 160, text="Donner le mois et l'année")
        self.bouton_MA = tk.Button( text="Soumettre le mois et l'année", command=lambda: self.AskMA()).place(x=20, y=220)
        self.boutonComptaAuto = tk.Button( text="Lancer compta automatique", command=lambda: self.TestCompta()).place(x=20, y=310)
        self.quitButon= tk.Button(self,text="Quitter", command=lambda :self.destroy()).place(x=330, y=360)

        self.canva1.create_text(30, 390, text=self.Compta.version)

    def AskComptaFile(self):

        FileC = filedialog.askopenfilename(initialdir="Desktop/", title="Feuille de compta")
        print(FileC)

        if ".xlsx" in FileC:
            if FileC != "":
                self.canva1.create_oval(360, 60, 390, 90, fill='green')
                return FileC
            else:
                self.canva1.create_oval(360, 60, 390, 90, fill='red')
                return ""
        else:
            self.AffichageMessage("ERREUR: Extension fichier non reconnue. \n Verifie le fichier que tu as donné au programme. \n Sinon appelle Maxime il va regler le probleme")
            self.canva1.create_oval(360, 60, 390, 90, fill='red')
            return ""


    def AskPackFile(self):
        FileP = filedialog.askdirectory(initialdir="Desktop/", title="Paquet de feuille de caisse")
        if FileP != "":
            self.canva1.create_oval(360,105,390,135, fill = 'green')  ## voyant check Paquet de copie
            return FileP
        else:
            self.canva1.create_oval(360,105,390,135, fill = 'red')  ## voyant check Paquet de copie
            return ""

    def AskMA(self):

        mois = self.variable.get()
        annee = self.variable2.get()

        self.Compta.AffilierMA(mois, annee)

        self.canva2.destroy()
        self.canva2 = tk.Canvas(self, width=200, height=75)
        self.canva2.create_text(100, 25, text= str(mois) + " " + str(annee), font=tkinter.font.Font(size=12))  ## Titre dans le rectangle
        self.canva2.place(x=200,y=225)

    def AffichageMessage(self,Message):
        newWindow = tk.Toplevel(app)
        newWindow.geometry('500x100')
        newWindow.resizable(width=0, height=0)
        newWindow.iconbitmap("Logo valid.ico")
        newWindow.title("Compta Probleme")
        label = tk.Label(newWindow, text=Message,).place(x=20,y=20)
        boutonok = tk.Button(newWindow, text="OK", command =lambda :newWindow.destroy()).place(x=135,y=65)

    def TestCompta(self):
        if (self.Compta.tablfile==""):
            self.AffichageMessage("Vous n'avez pas donner de tableur de compta")

        elif(self.Compta.paquetfile==""):
            self.AffichageMessage("Vous n'avez pas donner de Paquet de feuille de caisse")

        elif (self.Compta.mois == ""):
            self.AffichageMessage("Vous n'avez pas donner le mois ")

        elif (self.Compta.annee == ""):
            self.AffichageMessage("Vous n'avez pas donner l'année ")

        elif(self.Compta.tablfile!="" and self.Compta.paquetfile!= "" and self.Compta.mois!="" and self.Compta.annee != 0):
            self.Compta.Execution()
            self.AffichageMessage("Compta effectuée")
            print(" COMPTA FINI")

if __name__ == "__main__":
    app = Application()
    app.mainloop()

print("chaussure :) " )   ##j'ai retrouvé mes chaussures !

